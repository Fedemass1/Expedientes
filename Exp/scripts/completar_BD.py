# Importa tu modelo Expedientes
import os
import openpyxl
from Exp.models import Expedientes


def run(*args):
    # Obtén el directorio del script actual
    script_dir = os.path.dirname(__file__)

    # Construye la ruta al archivo usando una ruta relativa
    file_path = os.path.join(script_dir, "../expedientes.xlsx")

    # Carga el archivo usando openpyxl
    excel_dataframe = openpyxl.load_workbook(file_path)
    dataframe = excel_dataframe.active

    # Itera sobre las filas del archivo Excel
    for row in dataframe.iter_rows(min_row=2, max_row=dataframe.max_row, values_only=True):

        # Crea una instancia del modelo Expedientes
        expediente = Expedientes()

        # Asigna valores a los campos del modelo desde la fila del archivo Excel
        expediente.fecha = row[0]
        expediente.nro_exp = row[1]
        expediente.iniciador = row[2]
        expediente.objeto = row[3]
        expediente.nro_resol_rectorado = row[4]
        expediente.nro_resol_CS = row[5]
        expediente.observaciones = row[6]

        # Guarda la instancia en la base de datos
        expediente.save()
